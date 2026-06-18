from __future__ import annotations

import csv
import hashlib
import io
import re
import shutil
import zipfile
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook

REPO_ROOT = Path(__file__).resolve().parents[1]
ZIP_PATH = Path(r"C:\Users\USER\Downloads\mwtcb_gemini_copywriting_landbank_csv.zip")
XLSX_PATH = Path(r"C:\Users\USER\Downloads\mwtcb_copywriting_landbank.xlsx")

BASE_DIR = REPO_ROOT / "data" / "copywriting_landbank"
PRODUCT_DIR = BASE_DIR / "products" / "mwtcb"
IMPORT_VERSION = "2026-06-19_initial_landbank"
BATCH_DIR = PRODUCT_DIR / "imports" / IMPORT_VERSION
RAW_INPUT_DIR = BATCH_DIR / "raw_input"
NORMALIZED_DIR = BATCH_DIR / "normalized"
EXPORTS_DIR = BASE_DIR / "exports"
SCHEMA_DIR = BASE_DIR / "schema"
GLOBAL_DIR = BASE_DIR / "global"

PRODUCT_ID = "MWTCB_25ML"
PRODUCT_NAME = "Minyak Warisan Tok Cap Burung"

SOURCE_BATCHES = {
    "xlsx": {
        "source_batch": "2026-06-19_xlsx",
        "source_file": XLSX_PATH.name,
    },
    "zip_csv": {
        "source_batch": "2026-06-19_zip_csv",
        "source_file": ZIP_PATH.name,
    },
}

RANKING_GROUP_MAP = {
    "Top_15_Expand": "top_expand_first",
    "Top 15 Expand": "top_expand_first",
    "Bundle_Top10": "top_bundle_offers",
    "Bundle Top10": "top_bundle_offers",
    "UGC_Top10": "top_ugc_scripts",
    "UGC Top10": "top_ugc_scripts",
    "ProductOnly_Top10": "top_product_only",
    "ProductOnly Top10": "top_product_only",
    "Poster_Top10": "top_poster_ads",
    "Poster Top10": "top_poster_ads",
    "Aggressive_Hooks": "top_aggressive_hooks",
    "Aggressive Hooks": "top_aggressive_hooks",
    "Safe_Hooks": "top_safe_hooks",
    "Safe Hooks": "top_safe_hooks",
    "Park_Later": "park_later",
    "Park Later": "park_later",
}

RANKING_GROUP_DESCRIPTIONS = {
    "top_expand_first": "Highest-priority motivation rows to expand first into future angle banks.",
    "top_bundle_offers": "Rows with strong multi-buy or offer-stacking potential.",
    "top_ugc_scripts": "Rows most suitable for creator-voiced UGC talking-head or POV scripts.",
    "top_product_only": "Rows that can carry product-only or product-dominant ads without a talent anchor.",
    "top_poster_ads": "Rows strongest for static poster, product card, or poster-first layouts.",
    "top_aggressive_hooks": "High-pressure hook shortlist preserved at source intensity.",
    "top_safe_hooks": "Safer or softer hook shortlist that still keeps the core sales angle.",
    "park_later": "Rows intentionally deprioritized for later expansion rather than deleted.",
}

NEAR_DUPLICATE_CLUSTERS = [
    {
        "cluster_id": "ND-001",
        "theme": "Balik kampung / practical gift for mak ayah",
        "canonical_row_id": f"{PRODUCT_ID}__zip_csv__BM014",
        "canonical_reason": "Stronger seasonal urgency and sharper gifting hook for family-return travel.",
        "variant_row_ids": [f"{PRODUCT_ID}__xlsx__BM004", f"{PRODUCT_ID}__xlsx__BM022"],
        "decision": "kept",
        "parked_row_ids": [],
        "note": "All three stay imported. The ZIP row is the canonical Part 2 expansion anchor; the XLSX rows preserve broader gifting variants.",
    },
    {
        "cluster_id": "ND-002",
        "theme": "Household drawer / standby location problem",
        "canonical_row_id": f"{PRODUCT_ID}__zip_csv__BM007",
        "canonical_reason": "Higher tension because it dramatizes the emergency search moment and fixed-location logic.",
        "variant_row_ids": [f"{PRODUCT_ID}__xlsx__BM006"],
        "decision": "kept",
        "parked_row_ids": [],
        "note": "Keep both. XLSX BM006 is the home-organisation slap angle; ZIP BM_007 is the panic-at-use angle.",
    },
    {
        "cluster_id": "ND-003",
        "theme": "Standby rumah / preparedness gap",
        "canonical_row_id": f"{PRODUCT_ID}__zip_csv__BM023",
        "canonical_reason": "More aggressive household inadequacy framing and clearer midnight-emergency visual.",
        "variant_row_ids": [f"{PRODUCT_ID}__xlsx__BM028"],
        "decision": "kept",
        "parked_row_ids": [],
        "note": "Keep both. The ZIP row is the sharper conversion-first anchor; the XLSX row fits cleaner rack-setup or poster language.",
    },
]

BUCKET_DESCRIPTIONS = {
    "FAMILY_HOME": "Family-duty, home-care, and parent-facing purchase situations.",
    "MULTI_BUY": "Angles that justify buying more than one bottle for spread placement or backup.",
    "NOSTALGIA_TRUST": "Old-school familiarity, heritage proof, and memory-triggered trust lanes.",
    "PERSONA_SPECIFIC": "Rows anchored to a tighter persona or occupational use case.",
    "PRACTICAL_STORAGE": "Small-bottle convenience, storage logic, and physical placement proof.",
    "SEASONAL_CONTEXT": "Campaigns linked to balik kampung, kenduri, open house, or weather context.",
    "STANDBY_BEFORE_NEED": "Preparedness lanes that sell before the discomfort hits.",
    "TIKTOK_CURIOSITY": "Scroll-stop or curiosity hooks designed to earn the next second of attention.",
    "TRAVEL_CAR_BAG": "Mobility, glovebox, roadtrip, travel bag, or on-the-go carry lanes.",
    "VISUAL_RECOGNITION": "Angles that sell through the bottle look, cap, color, or heritage silhouette.",
}

PERSONA_DESCRIPTIONS = {
    "anak_beli_untuk_parent": "Child-to-parent gifting or practical-care buyer.",
    "driver_travel": "Driver, commuter, or travel-heavy buyer who needs an in-car standby item.",
    "family_home_buyer": "Home-oriented family purchaser buying for household completeness.",
    "impulse_tiktok_buyer": "Low-friction social-commerce buyer responding to offer pressure or quick proof.",
    "lelaki_bekerja": "Working male or labor-oriented buyer with strain, fatigue, or carry-pocket context.",
    "mak_mak": "Mother-led household buyer with urgency around children or home readiness.",
    "nostalgia_buyer": "Buyer pulled primarily by heritage look, memory, or old-household trust.",
    "universal_household": "Broad household persona that can fit product-only, poster, or video entry points.",
}

BOLDNESS_DESCRIPTIONS = {
    "AGGRESSIVE": "High-pressure, blunt, or guilt-forward direct-response language.",
    "BOLD": "Clear friction and urgency without going fully hard-slap.",
    "MODERATE": "Balanced persuasion with emotional or practical proof.",
    "SOFT": "Lowest-intensity variant that still preserves the core sales lane.",
}

USAGE_CONTEXT_DESCRIPTIONS = {
    "balik_kampung": "Travel or gifting context tied to returning home to family.",
    "beg": "Bag or handbag carry logic.",
    "family_gathering": "Family visit, shared-household, or communal context.",
    "kereta": "Car or glovebox placement context.",
    "kerja": "Work, commute, or task-fatigue context.",
    "laci": "Drawer, cabinet, or fixed household-storage context.",
    "malam": "Night-time or late-hour urgency context.",
    "open_house": "Open house or festive hosting context.",
    "rak": "Shelf or visual household-readiness context.",
    "rumah": "General in-home use or household readiness context.",
    "travel": "Trip, packing, or movement-heavy context beyond the car.",
}

SALES_MECHANISM_DESCRIPTIONS = {
    "before_need_preparation": "Sell preparedness before the pain or discomfort arrives.",
    "family_responsibility": "Sell through guilt, duty, or care toward family members.",
    "fear_of_not_having": "Sell through the downside of not keeping a standby bottle.",
    "gift_utility": "Sell practical gifting rather than decorative gifting.",
    "household_readiness": "Sell the idea of a complete, well-prepared home.",
    "impulse_low_friction": "Sell through low price, quick checkout, or immediate cart logic.",
    "multi_location_logic": "Sell one bottle per location so the product is always nearby.",
    "nostalgia_trust": "Sell through heritage proof and memory-based familiarity.",
    "practical_convenience": "Sell size, portability, and storage simplicity.",
    "visual_recognition": "Sell via instantly recognizable product visuals.",
}

CONTENT_FORMAT_DESCRIPTIONS = {
    "POV_scene": "Scenario-first POV or acted moment designed for short-form motion.",
    "UGC_talking_head": "Creator-led direct talk-to-camera delivery.",
    "carousel": "Slide-based static sequence or still-led product breakdown.",
    "hybrid_video": "Mixed product proof and talent scene video lane.",
    "poster_ad": "Static poster or sales-card layout lane.",
    "product_only_video": "Product-dominant motion without a strong human anchor.",
}

PLATFORM_SURFACE_DESCRIPTIONS = {
    "Shopee_Lazada_visual": "Marketplace visual surface such as PDP images or listing cards.",
    "TikTok_Shop_video": "TikTok Shop short-form video surface.",
    "TikTok_live_pin": "Live-stream pin or live-support surface.",
    "WhatsApp_blast": "WhatsApp resale or broadcast collateral lane.",
    "poster_ads": "Poster, static ad, or boosted-image surface.",
    "product_page_asset": "Static or embedded product page asset slot.",
}

FORMULA_DESCRIPTIONS = {
    "4U": "Urgency-forward direct-response pattern used for practical conversion hooks.",
    "AIDA": "Attention, Interest, Desire, Action pattern.",
    "BAB": "Before, After, Bridge transformation pattern.",
    "FAB": "Features, Advantages, Benefits pattern.",
    "HSO": "Hook, Story, Offer pattern.",
    "Hook/Family": "Family-duty hook pattern preserved from source wording.",
    "Hook/Family/Product/CTA": "Family-duty hook that moves quickly into product and CTA.",
    "Hook/Fear of Not Having": "Preparedness hook built on the downside of having no standby item.",
    "Hook/Memory": "Memory trigger hook centered on recognition and nostalgia.",
    "Hook/Memory/Trust/CTA": "Recognition-led hook that leans into trust proof then CTA.",
    "Hook/Multi-location": "Location-spread logic that sells multi-buy or placement convenience.",
    "Hook/Multi-location/CTA": "Location-spread logic with explicit close.",
    "Hook/Pain": "Direct pain-first hook.",
    "Hook/Use Moment": "Moment-of-use opening that starts inside the scenario.",
    "Hook/Use Moment/CTA": "Moment-of-use opening that closes fast.",
    "Hook/Visual Recognition": "Product-look or packaging-driven recognition hook.",
    "Hook/Visual/Reason/CTA": "Visual proof then buying reason then CTA.",
    "PAS": "Problem, Agitate, Solution pattern.",
    "PASTOR": "Problem, Amplify, Story, Transformation, Offer, Response pattern.",
    "PPPP": "Picture, Promise, Prove, Push pattern.",
    "Problem/Moment": "Problem-first moment snapshot.",
    "Problem/Moment/Standby/CTA": "Problem snapshot followed by standby logic and CTA.",
    "QUEST": "Qualify, Understand, Educate, Stimulate, Transition pattern.",
    "SLAP": "Short, direct slap-style hook pattern preserved from source framing.",
}


def ensure_dirs() -> None:
    for path in [
        BASE_DIR,
        PRODUCT_DIR,
        BATCH_DIR,
        RAW_INPUT_DIR,
        NORMALIZED_DIR,
        EXPORTS_DIR,
        SCHEMA_DIR,
        GLOBAL_DIR,
    ]:
        path.mkdir(parents=True, exist_ok=True)


def snake_case(value: str) -> str:
    value = "" if value is None else str(value).strip()
    value = re.sub(r"(?<!^)(?=[A-Z])", "_", value)
    value = re.sub(r"[^A-Za-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value.lower()


def normalize_source_id(raw_value: str) -> str:
    raw_value = (raw_value or "").strip().upper().replace(" ", "")
    match = re.match(r"BM_?(\d+)$", raw_value)
    if match:
        return f"BM{int(match.group(1)):03d}"
    return raw_value.replace("_", "")


def sha256_for(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_yaml(path: Path, payload: object) -> None:
    with path.open("w", encoding="utf-8") as handle:
        yaml.safe_dump(payload, handle, sort_keys=False, allow_unicode=True, width=1000)


def snapshot_raw_inputs() -> tuple[dict[str, list[dict[str, str]]], dict[str, list[dict[str, str]]]]:
    shutil.copy2(XLSX_PATH, RAW_INPUT_DIR / XLSX_PATH.name)
    shutil.copy2(ZIP_PATH, RAW_INPUT_DIR / ZIP_PATH.name)

    zip_tables: dict[str, list[dict[str, str]]] = {}
    with zipfile.ZipFile(ZIP_PATH) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            data = archive.read(info.filename)
            (RAW_INPUT_DIR / f"zip__{Path(info.filename).name}").write_bytes(data)
            if info.filename.lower().endswith(".csv"):
                text = data.decode("utf-8-sig", errors="replace")
                zip_tables[Path(info.filename).stem] = list(csv.DictReader(io.StringIO(text)))

    workbook = load_workbook(XLSX_PATH, data_only=True)
    xlsx_tables: dict[str, list[dict[str, str]]] = {}
    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        data_rows: list[dict[str, str]] = []
        for row in rows[1:]:
            record: dict[str, str] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = "" if index >= len(row) or row[index] is None else str(row[index]).strip()
                record[header] = value
            if any(value != "" for value in record.values()):
                data_rows.append(record)
        xlsx_tables[worksheet.title] = data_rows
        csv_target = RAW_INPUT_DIR / f"xlsx__{snake_case(worksheet.title)}.csv"
        with csv_target.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for record in data_rows:
                writer.writerow([record.get(header, "") for header in headers])

    return xlsx_tables, zip_tables


def ingest_core_rows(
    rows: list[dict[str, str]],
    kind: str,
    batch_key: str,
    source_sheet: str,
    canonical_lookup: dict[tuple[str, str], str],
    buyer_rows: list[dict[str, str]],
    classification_rows: list[dict[str, str]],
    normalized_snapshots: list[tuple[Path, list[dict[str, str]]]],
) -> None:
    config = SOURCE_BATCHES[batch_key]
    normalized_rows: list[dict[str, str]] = []
    for index, row in enumerate(rows, start=1):
        normalized = {snake_case(key): ("" if value is None else str(value).strip()) for key, value in row.items() if key is not None}
        source_id_raw = normalized.get("motivation_id", "")
        source_id = normalize_source_id(source_id_raw)
        motivation_batch_key = f"{config['source_batch']}__{source_id}"
        if kind == "buyer":
            row_id = f"{PRODUCT_ID}__{batch_key}__{source_id}"
            ordered = {
                "buyer_motivation_row_id": row_id,
                "product_id": PRODUCT_ID,
                "source_motivation_id": source_id,
                "source_motivation_id_raw": source_id_raw,
                "motivation_batch_key": motivation_batch_key,
                "source_file": config["source_file"],
                "source_sheet": source_sheet,
                "source_batch": config["source_batch"],
                "import_version": IMPORT_VERSION,
                "source_sequence": index,
                "buyer_situation": normalized.get("buyer_situation", ""),
                "buyer_problem": normalized.get("buyer_problem", ""),
                "buyer_desire": normalized.get("buyer_desire", ""),
                "what_buyer_is_really_thinking": normalized.get("what_buyer_is_really_thinking", ""),
                "emotional_trigger": normalized.get("emotional_trigger", ""),
                "practical_trigger": normalized.get("practical_trigger", ""),
                "buying_trigger": normalized.get("buying_trigger", ""),
                "product_role": normalized.get("product_role", ""),
                "proof_from_product": normalized.get("proof_from_product", ""),
                "best_formula": normalized.get("best_formula", ""),
                "hook_direction": normalized.get("hook_direction", ""),
                "sample_hook": normalized.get("sample_hook", ""),
                "sample_subhook": normalized.get("sample_subhook", ""),
                "sample_cta": normalized.get("sample_cta", ""),
                "why_it_can_sell": normalized.get("why_it_can_sell", ""),
            }
            canonical_lookup[(config["source_batch"], source_id)] = row_id
            buyer_rows.append(ordered)
            normalized_rows.append(ordered)
        else:
            row_id = f"{PRODUCT_ID}__{batch_key}__{source_id}"
            ordered = {
                "motivation_classification_row_id": row_id,
                "product_id": PRODUCT_ID,
                "source_motivation_id": source_id,
                "source_motivation_id_raw": source_id_raw,
                "motivation_batch_key": motivation_batch_key,
                "buyer_motivation_row_id": f"{PRODUCT_ID}__{batch_key}__{source_id}",
                "source_file": config["source_file"],
                "source_sheet": source_sheet,
                "source_batch": config["source_batch"],
                "import_version": IMPORT_VERSION,
                "source_sequence": index,
                "primary_bucket": normalized.get("primary_bucket", ""),
                "secondary_bucket": normalized.get("secondary_bucket", ""),
                "buyer_stage": normalized.get("buyer_stage", ""),
                "persona_fit": normalized.get("persona_fit", ""),
                "boldness_level": normalized.get("boldness_level", ""),
                "usage_context": normalized.get("usage_context", ""),
                "sales_mechanism": normalized.get("sales_mechanism", ""),
                "best_content_format": normalized.get("best_content_format", ""),
                "best_platform_surface": normalized.get("best_platform_surface", ""),
                "bundle_potential": normalized.get("bundle_potential", ""),
                "repeat_purchase_potential": normalized.get("repeat_purchase_potential", ""),
                "strongest_hook_from_this_motivation": normalized.get("strongest_hook_from_this_motivation", ""),
                "expansion_priority": normalized.get("expansion_priority", ""),
                "notes_for_angle_generation": normalized.get("notes_for_angle_generation", ""),
            }
            classification_rows.append(ordered)
            normalized_rows.append(ordered)
    filename = "buyer_motivations" if kind == "buyer" else "motivation_classification"
    normalized_snapshots.append((NORMALIZED_DIR / f"{filename}__{batch_key}.csv", normalized_rows))


def dedupe_rows(
    buyer_rows: list[dict[str, str]],
    classification_rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    buyer_deduped: list[dict[str, str]] = []
    seen_buyer_keys: dict[tuple[str, ...], str] = {}
    exact_buyer_duplicates: list[dict[str, str]] = []
    buyer_key_fields = [
        "buyer_situation",
        "buyer_problem",
        "buyer_desire",
        "what_buyer_is_really_thinking",
        "emotional_trigger",
        "practical_trigger",
        "buying_trigger",
        "product_role",
        "proof_from_product",
        "best_formula",
        "hook_direction",
        "sample_hook",
        "sample_subhook",
        "sample_cta",
        "why_it_can_sell",
    ]
    for row in buyer_rows:
        key = tuple(row[field].strip() for field in buyer_key_fields)
        if key in seen_buyer_keys:
            exact_buyer_duplicates.append(
                {
                    "removed_row_id": row["buyer_motivation_row_id"],
                    "kept_row_id": seen_buyer_keys[key],
                }
            )
            continue
        seen_buyer_keys[key] = row["buyer_motivation_row_id"]
        buyer_deduped.append(row)

    class_deduped: list[dict[str, str]] = []
    seen_class_keys: dict[tuple[str, ...], str] = {}
    exact_class_duplicates: list[dict[str, str]] = []
    class_key_fields = [
        "buyer_motivation_row_id",
        "primary_bucket",
        "secondary_bucket",
        "buyer_stage",
        "persona_fit",
        "boldness_level",
        "usage_context",
        "sales_mechanism",
        "best_content_format",
        "best_platform_surface",
        "bundle_potential",
        "repeat_purchase_potential",
        "strongest_hook_from_this_motivation",
        "expansion_priority",
        "notes_for_angle_generation",
    ]
    for row in classification_rows:
        key = tuple(row[field].strip() for field in class_key_fields)
        if key in seen_class_keys:
            exact_class_duplicates.append(
                {
                    "removed_row_id": row["motivation_classification_row_id"],
                    "kept_row_id": seen_class_keys[key],
                }
            )
            continue
        seen_class_keys[key] = row["motivation_classification_row_id"]
        class_deduped.append(row)

    return buyer_deduped, class_deduped, exact_buyer_duplicates, exact_class_duplicates


def build_rankings(
    xlsx_tables: dict[str, list[dict[str, str]]],
    zip_tables: dict[str, list[dict[str, str]]],
    canonical_lookup: dict[tuple[str, str], str],
) -> tuple[dict[str, object], list[tuple[Path, list[dict[str, str]], list[str]]]]:
    payload: dict[str, object] = {
        "product_id": PRODUCT_ID,
        "product_name": PRODUCT_NAME,
        "import_version": IMPORT_VERSION,
        "ranking_groups": {},
    }
    normalized_csvs: list[tuple[Path, list[dict[str, str]], list[str]]] = []
    rank_sheet_order = [
        ("xlsx", "Top 15 Expand"),
        ("zip_csv", "Top_15_Expand"),
        ("xlsx", "Bundle Top10"),
        ("zip_csv", "Bundle_Top10"),
        ("xlsx", "UGC Top10"),
        ("zip_csv", "UGC_Top10"),
        ("xlsx", "ProductOnly Top10"),
        ("zip_csv", "ProductOnly_Top10"),
        ("xlsx", "Poster Top10"),
        ("zip_csv", "Poster_Top10"),
        ("xlsx", "Aggressive Hooks"),
        ("zip_csv", "Aggressive_Hooks"),
        ("xlsx", "Safe Hooks"),
        ("zip_csv", "Safe_Hooks"),
        ("xlsx", "Park Later"),
        ("zip_csv", "Park_Later"),
    ]
    for batch_key, table_name in rank_sheet_order:
        rows = xlsx_tables.get(table_name, []) if batch_key == "xlsx" else zip_tables.get(table_name, [])
        if not rows:
            continue
        group_id = RANKING_GROUP_MAP[table_name]
        config = SOURCE_BATCHES[batch_key]
        fieldnames: list[str] = []
        normalized_entries: list[dict[str, str]] = []
        for index, row in enumerate(rows, start=1):
            normalized = {snake_case(key): ("" if value is None else str(value).strip()) for key, value in row.items() if key is not None}
            source_id = normalize_source_id(normalized.get("motivation_id", "")) if normalized.get("motivation_id", "") else ""
            entry = {
                "source_batch": config["source_batch"],
                "source_file": config["source_file"],
                "source_sheet": table_name,
                "import_version": IMPORT_VERSION,
                "source_sequence": index,
            }
            if source_id:
                entry["source_motivation_id"] = source_id
                entry["buyer_motivation_row_id"] = canonical_lookup.get((config["source_batch"], source_id), "")
            for key, value in normalized.items():
                if key == "motivation_id":
                    continue
                entry[key] = value
            normalized_entries.append(entry)
            for key in entry.keys():
                if key not in fieldnames:
                    fieldnames.append(key)
        ranking_groups = payload["ranking_groups"]
        assert isinstance(ranking_groups, dict)
        if group_id not in ranking_groups:
            ranking_groups[group_id] = {
                "description": RANKING_GROUP_DESCRIPTIONS[group_id],
                "source_lists": [],
            }
        ranking_groups[group_id]["source_lists"].append(
            {
                "source_batch": config["source_batch"],
                "source_file": config["source_file"],
                "source_sheet": table_name,
                "entry_count": len(normalized_entries),
                "entries": normalized_entries,
            }
        )
        normalized_csvs.append(
            (NORMALIZED_DIR / f"rankings__{group_id}__{batch_key}.csv", normalized_entries, fieldnames)
        )
    return payload, normalized_csvs


def write_taxonomies(buyer_rows: list[dict[str, str]], classification_rows: list[dict[str, str]]) -> None:
    formula_values = sorted({row["best_formula"] for row in buyer_rows if row["best_formula"]})
    bucket_values = sorted(
        {row["primary_bucket"] for row in classification_rows if row["primary_bucket"]}
        | {row["secondary_bucket"] for row in classification_rows if row["secondary_bucket"]}
    )
    persona_values = sorted({row["persona_fit"] for row in classification_rows if row["persona_fit"]})
    boldness_values = sorted({row["boldness_level"] for row in classification_rows if row["boldness_level"]})
    usage_context_values = sorted({row["usage_context"] for row in classification_rows if row["usage_context"]})
    sales_mechanism_values = sorted({row["sales_mechanism"] for row in classification_rows if row["sales_mechanism"]})
    content_format_values = sorted({row["best_content_format"] for row in classification_rows if row["best_content_format"]})
    platform_surface_values = sorted({row["best_platform_surface"] for row in classification_rows if row["best_platform_surface"]})

    write_yaml(
        GLOBAL_DIR / "copywriting_formulas.yaml",
        {
            "taxonomy_id": "copywriting_formulas",
            "version": 1,
            "description": "Reusable formula inventory observed in the seeded MWTCB landbank. New products may reuse or extend this list.",
            "values": [
                {"id": value, "description": FORMULA_DESCRIPTIONS.get(value, "Source-preserved formula label.")}
                for value in formula_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "bucket_taxonomy.yaml",
        {
            "taxonomy_id": "bucket_taxonomy",
            "version": 1,
            "description": "Commercial angle buckets used to classify buyer motivations.",
            "values": [
                {"id": value, "description": BUCKET_DESCRIPTIONS.get(value, "Source-derived bucket.")}
                for value in bucket_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "persona_taxonomy.yaml",
        {
            "taxonomy_id": "persona_taxonomy",
            "version": 1,
            "description": "Persona-fit values observed in classification rows and reusable for future products.",
            "values": [
                {"id": value, "description": PERSONA_DESCRIPTIONS.get(value, "Source-derived persona fit.")}
                for value in persona_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "boldness_taxonomy.yaml",
        {
            "taxonomy_id": "boldness_taxonomy",
            "version": 1,
            "description": "Copy intensity scale. Preserve source boldness labels; do not auto-soften aggressive rows during import.",
            "values": [
                {"id": value, "description": BOLDNESS_DESCRIPTIONS.get(value, "Source-derived boldness label.")}
                for value in boldness_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "usage_context_taxonomy.yaml",
        {
            "taxonomy_id": "usage_context_taxonomy",
            "version": 1,
            "description": "Context or placement cues used during angle expansion and creative routing.",
            "values": [
                {"id": value, "description": USAGE_CONTEXT_DESCRIPTIONS.get(value, "Source-derived usage context.")}
                for value in usage_context_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "sales_mechanism_taxonomy.yaml",
        {
            "taxonomy_id": "sales_mechanism_taxonomy",
            "version": 1,
            "description": "Primary commercial mechanism each motivation uses to move the buyer toward action.",
            "values": [
                {"id": value, "description": SALES_MECHANISM_DESCRIPTIONS.get(value, "Source-derived sales mechanism.")}
                for value in sales_mechanism_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "content_format_taxonomy.yaml",
        {
            "taxonomy_id": "content_format_taxonomy",
            "version": 1,
            "description": "Creative-format preferences inferred from classification rows.",
            "values": [
                {"id": value, "description": CONTENT_FORMAT_DESCRIPTIONS.get(value, "Source-derived content format.")}
                for value in content_format_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "platform_surface_taxonomy.yaml",
        {
            "taxonomy_id": "platform_surface_taxonomy",
            "version": 1,
            "description": "Output surfaces observed in the seed landbank and reusable for future products.",
            "values": [
                {"id": value, "description": PLATFORM_SURFACE_DESCRIPTIONS.get(value, "Source-derived platform surface.")}
                for value in platform_surface_values
            ],
        },
    )
    write_yaml(
        GLOBAL_DIR / "copy_asset_types.yaml",
        {
            "taxonomy_id": "copy_asset_types",
            "version": 1,
            "description": "Repository-native asset families supported by the copywriting landbank module.",
            "values": [
                {"id": "product_profile", "description": "Product-level copy authority and truth spine."},
                {"id": "buyer_motivation", "description": "Scenario-level buyer motivation rows."},
                {"id": "motivation_classification", "description": "Classification rows layered on buyer motivations."},
                {"id": "ranking_manifest", "description": "Top-list or park-later ranking groups."},
                {"id": "angle", "description": "Future Part 2 angle bank rows."},
                {"id": "hook", "description": "Future hook-bank rows."},
                {"id": "subhook", "description": "Future subhook-bank rows."},
                {"id": "usp", "description": "Future USP-bank rows."},
                {"id": "cta", "description": "Future CTA-bank rows."},
                {"id": "video_copy_matrix", "description": "Future video matrix rows linking angles, hooks, and formats."},
                {"id": "poster_copy_matrix", "description": "Future poster matrix rows linking angles, hooks, and layouts."},
                {"id": "import_report", "description": "Human-readable import trace for a product or batch."},
                {"id": "duplicate_report", "description": "Human-readable duplicate and near-duplicate audit."},
            ],
        },
    )


def write_schemas(buyer_fieldnames: list[str], class_fieldnames: list[str]) -> None:
    schemas = {
        "product_profile.schema.yaml": {
            "schema_id": "product_profile",
            "version": 1,
            "format": "yaml",
            "intended_usage": ["Product-level copy authority", "Downstream creative planning", "Future angle expansion anchor"],
            "required_fields": [
                {"name": "product_id", "type": "string"},
                {"name": "product_name", "type": "string"},
                {"name": "category", "type": "string"},
                {"name": "market", "type": "string"},
                {"name": "language_style", "type": "string or list"},
                {"name": "product_truth", "type": "list"},
                {"name": "visual_identity", "type": "list"},
                {"name": "trust_cues", "type": "list"},
                {"name": "core_selling_spine", "type": "mapping"},
                {"name": "recommended_platforms", "type": "list"},
            ],
            "optional_fields": [
                {"name": "source_scope", "type": "mapping"},
                {"name": "notes", "type": "list"},
            ],
        },
        "buyer_motivation.schema.yaml": {
            "schema_id": "buyer_motivation",
            "version": 1,
            "format": "csv",
            "primary_key": "buyer_motivation_row_id",
            "required_columns": buyer_fieldnames,
            "optional_columns": [],
            "intended_usage": ["Video copywriting", "Poster copywriting", "Future angle bank expansion", "Notion row generation"],
        },
        "motivation_classification.schema.yaml": {
            "schema_id": "motivation_classification",
            "version": 1,
            "format": "csv",
            "primary_key": "motivation_classification_row_id",
            "required_columns": class_fieldnames,
            "optional_columns": [],
            "intended_usage": ["Taxonomy filtering", "Ranking expansion", "Format routing"],
        },
        "angle.schema.yaml": {
            "schema_id": "angle",
            "version": 1,
            "format": "csv",
            "primary_key": "angle_row_id",
            "required_columns": [
                "angle_row_id",
                "product_id",
                "source_motivation_row_id",
                "angle_name",
                "angle_statement",
                "best_formula",
                "primary_bucket",
                "boldness_level",
            ],
            "optional_columns": ["hook_seed", "subhook_seed", "usp_seed", "cta_seed", "notes", "source_batch", "import_version"],
            "intended_usage": ["Part 2 Angle Master Bank", "Video and poster matrix seeding"],
        },
        "hook.schema.yaml": {
            "schema_id": "hook",
            "version": 1,
            "format": "csv",
            "primary_key": "hook_row_id",
            "required_columns": ["hook_row_id", "product_id", "hook_text", "boldness_level", "source_asset_type"],
            "optional_columns": ["source_motivation_row_id", "primary_bucket", "sales_mechanism", "notes"],
            "intended_usage": ["Standalone hook bank", "Creative testing lanes"],
        },
        "subhook.schema.yaml": {
            "schema_id": "subhook",
            "version": 1,
            "format": "csv",
            "primary_key": "subhook_row_id",
            "required_columns": ["subhook_row_id", "product_id", "subhook_text", "source_asset_type"],
            "optional_columns": ["source_motivation_row_id", "hook_row_id", "notes"],
            "intended_usage": ["Supporting copy bank", "Poster and video expansion"],
        },
        "usp.schema.yaml": {
            "schema_id": "usp",
            "version": 1,
            "format": "csv",
            "primary_key": "usp_row_id",
            "required_columns": ["usp_row_id", "product_id", "usp_text", "proof_type"],
            "optional_columns": ["source_motivation_row_id", "trust_cue", "notes"],
            "intended_usage": ["Proof-bank management", "Future angle-to-proof linking"],
        },
        "cta.schema.yaml": {
            "schema_id": "cta",
            "version": 1,
            "format": "csv",
            "primary_key": "cta_row_id",
            "required_columns": ["cta_row_id", "product_id", "cta_text", "boldness_level"],
            "optional_columns": ["source_motivation_row_id", "best_platform_surface", "notes"],
            "intended_usage": ["CTA bank", "Format-specific closing copy"],
        },
        "video_copy_matrix.schema.yaml": {
            "schema_id": "video_copy_matrix",
            "version": 1,
            "format": "csv",
            "primary_key": "video_matrix_row_id",
            "required_columns": ["video_matrix_row_id", "product_id", "angle_row_id", "hook_row_id", "content_format", "platform_surface"],
            "optional_columns": ["subhook_row_id", "usp_row_id", "cta_row_id", "persona_fit", "notes"],
            "intended_usage": ["TikTok Shop video planning", "UGC and hybrid script routing"],
        },
        "poster_copy_matrix.schema.yaml": {
            "schema_id": "poster_copy_matrix",
            "version": 1,
            "format": "csv",
            "primary_key": "poster_matrix_row_id",
            "required_columns": ["poster_matrix_row_id", "product_id", "angle_row_id", "hook_row_id", "platform_surface"],
            "optional_columns": ["subhook_row_id", "usp_row_id", "cta_row_id", "visual_priority_notes", "notes"],
            "intended_usage": ["Poster copy planning", "Static-card buildout", "Product-page asset generation"],
        },
    }
    for filename, payload in schemas.items():
        write_yaml(SCHEMA_DIR / filename, payload)


def write_product_files(
    xlsx_tables: dict[str, list[dict[str, str]]],
    zip_tables: dict[str, list[dict[str, str]]],
    buyer_rows: list[dict[str, str]],
    class_rows: list[dict[str, str]],
    rankings_payload: dict[str, object],
    exact_buyer_duplicates: list[dict[str, str]],
    exact_class_duplicates: list[dict[str, str]],
) -> None:
    product_profile = {
        "product_id": PRODUCT_ID,
        "product_name": PRODUCT_NAME,
        "category": "Traditional topical oil",
        "market": "Malaysia",
        "language_style": [
            "Bahasa Melayu direct-response",
            "household vernacular",
            "old-school trust cues",
        ],
        "product_truth": [
            "Botol kecil 25ml.",
            "Minyak hijau.",
            "Cap merah.",
            "Botol kaca hijau tebal / tradisi.",
            "Label warisan dengan logo burung.",
            "Trust cue source copy includes “Sejak 1958”, “Formula Asli sejak 1958”, and “Petua Turun Temurun 1958”.",
        ],
        "visual_identity": [
            "Siluet botol kecil yang mudah disimpan dalam laci, beg, dan kereta.",
            "Cap merah yang senang dicam semasa kecemasan atau carian pantas.",
            "Warna minyak hijau dan botol kaca hijau retro / old-school.",
            "Label warisan dan logo burung sebagai pemicu visual recognition.",
        ],
        "trust_cues": [
            "Sejak 1958.",
            "Formula Asli sejak 1958.",
            "Petua Turun Temurun 1958.",
            "Mak ayah / orang tua kenal dan biasa simpan.",
        ],
        "core_selling_spine": {
            "primary_lanes": [
                "Standby before need: rumah wajib ada sebelum badan atau anak mula buat hal.",
                "Family-home trust: barang kecil yang mak ayah kenal, guna, dan percaya.",
                "Practical storage and multi-location logic: rumah, kereta, beg, rak, laci.",
            ],
            "secondary_lanes": [
                "Travel / balik kampung / gifting utility.",
                "Visual recognition through botol hijau, cap merah, dan label warisan.",
                "Repeat-buy and bundle logic driven by small-bottle spread placement.",
            ],
        },
        "recommended_platforms": [
            "TikTok_Shop_video",
            "poster_ads",
            "product_page_asset",
            "TikTok_live_pin",
            "Shopee_Lazada_visual",
        ],
        "source_scope": {
            "import_version": IMPORT_VERSION,
            "source_batches": [config["source_batch"] for config in SOURCE_BATCHES.values()],
        },
    }
    write_yaml(PRODUCT_DIR / "product.copy.yaml", product_profile)

    source_manifest = {
        "product_id": PRODUCT_ID,
        "product_name": PRODUCT_NAME,
        "import_version": IMPORT_VERSION,
        "canonical_outputs": {
            "buyer_motivations_csv": "data/copywriting_landbank/products/mwtcb/buyer_motivations.csv",
            "motivation_classification_csv": "data/copywriting_landbank/products/mwtcb/motivation_classification.csv",
            "rankings_yaml": "data/copywriting_landbank/products/mwtcb/rankings.yaml",
            "product_profile_yaml": "data/copywriting_landbank/products/mwtcb/product.copy.yaml",
        },
        "source_files": [
            {
                "source_batch": SOURCE_BATCHES["xlsx"]["source_batch"],
                "source_file": XLSX_PATH.name,
                "sha256": sha256_for(XLSX_PATH),
                "size_bytes": XLSX_PATH.stat().st_size,
                "usable_tables": list(xlsx_tables.keys()),
            },
            {
                "source_batch": SOURCE_BATCHES["zip_csv"]["source_batch"],
                "source_file": ZIP_PATH.name,
                "sha256": sha256_for(ZIP_PATH),
                "size_bytes": ZIP_PATH.stat().st_size,
                "usable_tables": list(zip_tables.keys()),
            },
        ],
        "batch_snapshots": {
            "raw_input_dir": str(RAW_INPUT_DIR.relative_to(REPO_ROOT)).replace("\\", "/"),
            "normalized_dir": str(NORMALIZED_DIR.relative_to(REPO_ROOT)).replace("\\", "/"),
        },
        "generated_exports": {
            "xlsx": "data/copywriting_landbank/exports/mwtcb_copywriting_landbank.xlsx",
            "zip": "data/copywriting_landbank/exports/mwtcb_copywriting_landbank_csv.zip",
        },
    }
    write_yaml(PRODUCT_DIR / "source_manifest.yaml", source_manifest)
    write_yaml(PRODUCT_DIR / "rankings.yaml", rankings_payload)

    index_payload = {
        "module_id": "copywriting_landbank",
        "version": 1,
        "canonical_storage": {
            "row_heavy_data": "CSV",
            "schemas_taxonomies_manifests": "YAML",
            "reports_docs": "Markdown",
            "excel_status": "export_only",
        },
        "products": [
            {
                "product_id": PRODUCT_ID,
                "path": "data/copywriting_landbank/products/mwtcb",
                "canonical_files": [
                    "product.copy.yaml",
                    "buyer_motivations.csv",
                    "motivation_classification.csv",
                    "rankings.yaml",
                    "source_manifest.yaml",
                    "import_report.md",
                    "duplicate_report.md",
                ],
            }
        ],
        "exports": [
            "data/copywriting_landbank/exports/mwtcb_copywriting_landbank.xlsx",
            "data/copywriting_landbank/exports/mwtcb_copywriting_landbank_csv.zip",
        ],
    }
    write_yaml(BASE_DIR / "index.yaml", index_payload)

    import_report_text = f"""# MWTCB Copywriting Landbank Import Report

## Scope

Imported the uploaded MWTCB copywriting landbank inputs into the repository-native copywriting landbank module.

- Product ID: `{PRODUCT_ID}`
- Product name: `{PRODUCT_NAME}`
- Import version: `{IMPORT_VERSION}`
- Canonical storage: CSV for row data, YAML for schemas/manifests/taxonomies, Markdown for reports
- Excel status: export-only

## Source Intake

- `mwtcb_copywriting_landbank.xlsx`
  - Buyer Motivation Map: {len(xlsx_tables['Buyer Motivation Map'])} rows
  - Classification Layer: {len(xlsx_tables['Classification Layer'])} rows
  - Ranking sheets: {', '.join([name for name in xlsx_tables.keys() if name not in ['Buyer Motivation Map', 'Classification Layer', 'README', 'Dashboard']])}
- `mwtcb_gemini_copywriting_landbank_csv.zip`
  - Buyer_Motivation_Map.csv: {len(zip_tables['Buyer_Motivation_Map'])} rows
  - Classification_Layer.csv: {len(zip_tables['Classification_Layer'])} rows
  - Ranking CSVs: {', '.join([name for name in zip_tables.keys() if name not in ['Buyer_Motivation_Map', 'Classification_Layer', 'README', 'Dashboard', 'Raw_Gemini_Notes', 'Categories']])}

## Canonical Outputs

- `buyer_motivations.csv`: {len(buyer_rows)} rows
- `motivation_classification.csv`: {len(class_rows)} rows
- `rankings.yaml`: 8 ranking groups with source-preserved sublists
- `product.copy.yaml`: seeded from uploaded landbank facts only
- `source_manifest.yaml`: source hashes, batch snapshots, and export pointers

## Normalization Decisions

- Headers normalized to canonical `snake_case`.
- Added `product_id`, source metadata, and deterministic canonical row ids.
- Preserved source copy intensity, including aggressive hooks and raw vernacular phrasing.
- Preserved both source batches; no merged rank order was invented across the two inputs.
- Ranking groups remain source-preserved inside each canonical group so future Part 2 work can decide cross-source winners explicitly.

## Exact Dedupe Result

- Buyer motivation rows removed as exact duplicates: {len(exact_buyer_duplicates)}
- Classification rows removed as exact duplicates: {len(exact_class_duplicates)}

## Near-Duplicate Review

- Near-duplicate clusters documented: {len(NEAR_DUPLICATE_CLUSTERS)}
- Rows parked due to duplicate review: 0
- `park_later` source ranking lists were preserved as prioritization guidance, not treated as deletions.

## Recommended Part 2 Next Step

Build the Angle Master Bank by starting with the union of both `top_expand_first` lists, then expand around the shared high-pressure lanes:

1. `STANDBY_BEFORE_NEED`
2. `FAMILY_HOME`
3. `NOSTALGIA_TRUST`
4. `MULTI_BUY`
5. `PRACTICAL_STORAGE` / `TRAVEL_CAR_BAG`

Keep cross-source duplicates as sibling variants until angle-level testing proves one should be demoted.
"""

    cluster_blocks = []
    for cluster in NEAR_DUPLICATE_CLUSTERS:
        cluster_blocks.append(
            "\n".join(
                [
                    f"### {cluster['cluster_id']} — {cluster['theme']}",
                    f"- Canonical row: `{cluster['canonical_row_id']}`",
                    f"- Variants kept: {', '.join(f'`{row_id}`' for row_id in cluster['variant_row_ids'])}",
                    f"- Decision: `{cluster['decision']}`",
                    f"- Reason: {cluster['canonical_reason']}",
                    f"- Note: {cluster['note']}",
                    "",
                ]
            )
        )

    duplicate_report_text = f"""# MWTCB Copywriting Landbank Duplicate Report

## Exact Duplicates Removed

- Buyer motivation rows removed: {len(exact_buyer_duplicates)}
- Classification rows removed: {len(exact_class_duplicates)}

## Near Duplicates Detected

{''.join(cluster_blocks)}## Rows Kept

- All {len(buyer_rows)} buyer motivation rows were retained in canonical CSV storage because no exact duplicates existed.
- Near-duplicate rows were kept when they served a different selling surface, urgency pattern, or gifting-vs-standby nuance.

## Rows Parked

- Rows parked by duplicate review: 0
- Source-provided `park_later` ranking entries remain in `rankings.yaml` as future expansion guidance and were not deleted from canonical storage.

## Decision Rule Applied

- Exact duplicates only: delete.
- Similar but commercially distinct: keep both, nominate one canonical anchor for future Part 2 expansion, document the sibling variant here.
- No copy was softened or rewritten beyond header normalization, row ids, and source metadata fields.
"""

    (BASE_DIR / "README.md").write_text(
        """# Copywriting Landbank

The copywriting landbank is the repo-native storage layer for commercial copy assets by product. It keeps row-heavy copy in CSV, reusable rules and schemas in YAML, and import / duplicate audits in Markdown.

## Canonical Data

- Canonical row data lives under `data/copywriting_landbank/products/<product_key>/`.
- CSV is the source of truth for row-heavy assets such as buyer motivations and classification layers.
- YAML is the source of truth for schemas, product profiles, taxonomies, rankings, and manifests.
- Markdown is the source of truth for human-facing import and duplicate reports.
- Excel under `data/copywriting_landbank/exports/` is export-only.

## Why Excel Is Export-Only

Excel is easy to inspect but poor as a canonical diff surface. The repo stores deterministic text-first artifacts so changes are reviewable, mergeable, and script-validated.

## Add a New Product

1. Create `data/copywriting_landbank/products/<product_key>/`.
2. Seed `product.copy.yaml`, `source_manifest.yaml`, and only the CSVs backed by real imported data.
3. Reuse the shared taxonomies in `data/copywriting_landbank/global/` and the schemas in `data/copywriting_landbank/schema/`.
4. Add a batch snapshot under `imports/<date>_<batch_name>/`.
5. Run `python scripts/validate_copywriting_landbank.py`.

## Add a New Batch For an Existing Product

1. Copy the untouched sources into `imports/<date>_<batch_name>/raw_input/`.
2. Write normalized snapshots into `imports/<date>_<batch_name>/normalized/` with the same metadata columns used in the canonical CSVs.
3. Merge only exact-new rows into the canonical CSVs.
4. Update `rankings.yaml`, `source_manifest.yaml`, `import_report.md`, and `duplicate_report.md`.
5. Regenerate the export artifacts.

## Part 2 Angle Master Bank Expansion

Do not create empty angle, hook, subhook, USP, CTA, video matrix, or poster matrix CSVs. Define the schemas now, then add those CSVs only when real data exists. Start Part 2 from `rankings.yaml` plus the near-duplicate canonical anchors documented in the duplicate report.

## Video Copywriting Use

- Use `buyer_motivations.csv` for raw scenario pain, desire, hook, subhook, CTA, and proof seeds.
- Use `motivation_classification.csv` to filter by persona, boldness, format, platform surface, and bucket.
- Use `rankings.yaml` to decide which rows to expand first for TikTok Shop, UGC, hybrid, and product-only lanes.

## Poster Copywriting Use

- Use the same buyer-motivation spine, but bias toward rows tagged `poster_ad`, `poster_ads`, `product_page_asset`, or `VISUAL_RECOGNITION` / `STANDBY_BEFORE_NEED` buckets.
- Preserve the strongest source hooks for poster headlines; do not auto-sanitize during import.

## Dedupe Rule

- Delete exact duplicates only.
- For near duplicates, keep both when the sales use is materially different and record the canonical-vs-variant decision in the duplicate report.
""",
        encoding="utf-8",
    )

    (PRODUCT_DIR / "README.md").write_text(
        f"""# MWTCB Copywriting Landbank

This folder contains the repository-native copywriting landbank for `{PRODUCT_NAME}` (`{PRODUCT_ID}`).

## Canonical Files

- `product.copy.yaml`: product-level copy profile from uploaded landbank facts only.
- `buyer_motivations.csv`: merged buyer motivation rows from both uploaded sources.
- `motivation_classification.csv`: merged classification rows aligned to the buyer rows.
- `rankings.yaml`: source-preserved ranking groups and park-later lists.
- `source_manifest.yaml`: import batches, hashes, and export pointers.
- `import_report.md`: batch summary and normalization trace.
- `duplicate_report.md`: exact and near-duplicate audit.

## Excel Status

- `data/copywriting_landbank/exports/mwtcb_copywriting_landbank.xlsx` is export-only.
- `data/copywriting_landbank/exports/mwtcb_copywriting_landbank_csv.zip` is export-only.
- Raw uploaded files are preserved under `imports/{IMPORT_VERSION}/raw_input/` for traceability.

## Add Another MWTCB Batch

1. Create a new folder under `imports/`.
2. Copy the untouched sources into `raw_input/`.
3. Normalize them into `normalized/` with the same metadata columns used in the canonical CSVs.
4. Merge exact-new rows only; document cross-batch overlaps in `duplicate_report.md`.
5. Regenerate exports and rerun validation.

## Future Part 2 Angle Master Bank

Use `rankings.yaml` plus the canonical near-duplicate anchors from `duplicate_report.md`. Start with `top_expand_first`, then branch into angle rows only after the scenario-level motivation is locked.

## Video and Poster Use

- Video: filter `motivation_classification.csv` by `best_content_format`, `best_platform_surface`, `persona_fit`, and `boldness_level`.
- Poster: prioritize headline-capable rows from `top_poster_ads`, `top_aggressive_hooks`, `top_safe_hooks`, and `VISUAL_RECOGNITION` / `STANDBY_BEFORE_NEED` buckets.

## Dedupe

Exact duplicates are removed. Near-duplicate rows stay when they provide different commercial leverage; the stronger expansion anchor is marked in the duplicate report.
""",
        encoding="utf-8",
    )

    for target in [
        PRODUCT_DIR / "import_report.md",
        BATCH_DIR / "import_report.md",
    ]:
        target.write_text(import_report_text, encoding="utf-8")
    for target in [
        PRODUCT_DIR / "duplicate_report.md",
        BATCH_DIR / "duplicate_report.md",
    ]:
        target.write_text(duplicate_report_text, encoding="utf-8")


def append_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    worksheet = workbook.create_sheet(title=title[:31])
    if not rows:
        worksheet.append(["empty"])
        return
    fieldnames = list(rows[0].keys())
    worksheet.append(fieldnames)
    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])


def write_exports(
    buyer_rows: list[dict[str, str]],
    class_rows: list[dict[str, str]],
    rankings_payload: dict[str, object],
) -> None:
    export_zip_path = EXPORTS_DIR / "mwtcb_copywriting_landbank_csv.zip"
    with zipfile.ZipFile(export_zip_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(PRODUCT_DIR / "buyer_motivations.csv", arcname="buyer_motivations.csv")
        archive.write(PRODUCT_DIR / "motivation_classification.csv", arcname="motivation_classification.csv")
        ranking_groups = rankings_payload["ranking_groups"]
        assert isinstance(ranking_groups, dict)
        for group_id, payload in ranking_groups.items():
            source_lists = payload["source_lists"]
            for source_list in source_lists:
                batch_slug = source_list["source_batch"].replace("2026-06-19_", "")
                csv_name = f"rankings__{group_id}__{batch_slug}.csv"
                rows = source_list["entries"]
                if rows:
                    buffer = io.StringIO()
                    fieldnames = list(rows[0].keys())
                    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(rows)
                    archive.writestr(csv_name, buffer.getvalue().encode("utf-8-sig"))

    export_xlsx_path = EXPORTS_DIR / "mwtcb_copywriting_landbank.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "README"
    for row in [
        ["Field", "Value"],
        ["product_id", PRODUCT_ID],
        ["product_name", PRODUCT_NAME],
        ["import_version", IMPORT_VERSION],
        ["buyer_motivation_rows", len(buyer_rows)],
        ["classification_rows", len(class_rows)],
        ["ranking_groups", len(rankings_payload["ranking_groups"])],
        ["excel_status", "export_only"],
    ]:
        worksheet.append(row)
    append_sheet(workbook, "buyer_motivations", buyer_rows)
    append_sheet(workbook, "motivation_class", class_rows)
    ranking_groups = rankings_payload["ranking_groups"]
    assert isinstance(ranking_groups, dict)
    for group_id, payload in ranking_groups.items():
        for source_list in payload["source_lists"]:
            batch_slug = "xlsx" if source_list["source_batch"].endswith("xlsx") else "zip"
            append_sheet(workbook, f"{group_id[:18]}_{batch_slug}", source_list["entries"])
    workbook.save(export_xlsx_path)


def main() -> None:
    ensure_dirs()
    xlsx_tables, zip_tables = snapshot_raw_inputs()

    canonical_lookup: dict[tuple[str, str], str] = {}
    buyer_rows: list[dict[str, str]] = []
    class_rows: list[dict[str, str]] = []
    normalized_snapshots: list[tuple[Path, list[dict[str, str]]]] = []

    ingest_core_rows(
        xlsx_tables["Buyer Motivation Map"],
        "buyer",
        "xlsx",
        "Buyer Motivation Map",
        canonical_lookup,
        buyer_rows,
        class_rows,
        normalized_snapshots,
    )
    ingest_core_rows(
        zip_tables["Buyer_Motivation_Map"],
        "buyer",
        "zip_csv",
        "Buyer_Motivation_Map",
        canonical_lookup,
        buyer_rows,
        class_rows,
        normalized_snapshots,
    )
    ingest_core_rows(
        xlsx_tables["Classification Layer"],
        "classification",
        "xlsx",
        "Classification Layer",
        canonical_lookup,
        buyer_rows,
        class_rows,
        normalized_snapshots,
    )
    ingest_core_rows(
        zip_tables["Classification_Layer"],
        "classification",
        "zip_csv",
        "Classification_Layer",
        canonical_lookup,
        buyer_rows,
        class_rows,
        normalized_snapshots,
    )

    buyer_deduped, class_deduped, exact_buyer_duplicates, exact_class_duplicates = dedupe_rows(
        buyer_rows, class_rows
    )
    buyer_fieldnames = list(buyer_deduped[0].keys())
    class_fieldnames = list(class_deduped[0].keys())
    write_csv(PRODUCT_DIR / "buyer_motivations.csv", buyer_deduped, buyer_fieldnames)
    write_csv(PRODUCT_DIR / "motivation_classification.csv", class_deduped, class_fieldnames)
    for path, rows in normalized_snapshots:
        if rows:
            write_csv(path, rows, list(rows[0].keys()))

    rankings_payload, normalized_ranking_csvs = build_rankings(
        xlsx_tables,
        zip_tables,
        canonical_lookup,
    )
    for path, rows, fieldnames in normalized_ranking_csvs:
        if rows:
            write_csv(path, rows, fieldnames)

    write_taxonomies(buyer_deduped, class_deduped)
    write_schemas(buyer_fieldnames, class_fieldnames)
    write_product_files(
        xlsx_tables,
        zip_tables,
        buyer_deduped,
        class_deduped,
        rankings_payload,
        exact_buyer_duplicates,
        exact_class_duplicates,
    )
    write_exports(buyer_deduped, class_deduped, rankings_payload)

    print(
        {
            "buyer_rows": len(buyer_deduped),
            "classification_rows": len(class_deduped),
            "exact_buyer_duplicates_removed": len(exact_buyer_duplicates),
            "exact_class_duplicates_removed": len(exact_class_duplicates),
            "near_duplicate_clusters": len(NEAR_DUPLICATE_CLUSTERS),
        }
    )


if __name__ == "__main__":
    main()
